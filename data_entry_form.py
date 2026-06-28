from pathlib import Path
from tkinter import END, Entry, Label, Tk, Button, Frame
from openpyxl import Workbook, load_workbook

#Define the path to the Excel file
EXCEL_FILE = Path(__file__).with_name("MDF-Data-Form.xlsx")
#Create global variables for the entry fields, status label, workbook, and worksheet
first_name_entry = None
last_name_entry = None
age_entry = None
email_entry = None
phone_entry = None
status_label = None
wb = None
ws = None

#Initialize the Excel workbook and worksheet
def init_excel():
    global wb, ws

    if EXCEL_FILE.exists():
        wb = load_workbook(EXCEL_FILE)
    else:
        wb = Workbook()

    ws = wb.active
    headers = ["First Name", "Last Name", "Age", "Email", "Phone"]

    if ws.max_row == 0 or ws.cell(row=1, column=1).value != headers[0]:
        for i, header in enumerate(headers, start=1):
            ws.cell(row=1, column=i).value = header

    wb.save(EXCEL_FILE)
    return wb, ws

#Create a function to add data to the Excel file
def add_data():
    first_name = first_name_entry.get().strip()
    last_name = last_name_entry.get().strip()
    age = age_entry.get().strip()
    email = email_entry.get().strip()
    phone = phone_entry.get().strip()

    if first_name and last_name and age and email and phone:
        ws.append([first_name, last_name, age, email, phone])
        wb.save(EXCEL_FILE)
        first_name_entry.delete(0, END)
        last_name_entry.delete(0, END)
        age_entry.delete(0, END)
        email_entry.delete(0, END)
        phone_entry.delete(0, END)
        status_label.config(text="Data added successfully!", fg="green")
    else:
        status_label.config(text="Please fill in all fields.", fg="red")

#Clearing the form fields and updating the status label
def clear_form():
    first_name_entry.delete(0, END)
    last_name_entry.delete(0, END)
    age_entry.delete(0, END)
    email_entry.delete(0, END)
    phone_entry.delete(0, END)
    status_label.config(text="Form cleared.", fg="blue")

#Validate the age input to ensure it is a number and not more than 3 digits
def validate_age(new_value):
    if new_value == "":
        return True
    return new_value.isdigit() and len(new_value) <= 3

#Move focus to the next field when pressing Enter
def next_field(event):
    event.widget.tk_focusNext().focus()

#Build the GUI for the data entry form
def build_gui():
    global first_name_entry, last_name_entry, age_entry, email_entry, phone_entry, status_label

    root = Tk()
    root.title("Data Entry Form")
    root.geometry("420x260")
    root.columnconfigure(1, weight=1)

    labels = ["First Name", "Last Name", "Age", "Email", "Phone"]
    entries = []
    #Enter the labels and entry fields for each data point, and bind the Enter key to move to the next field
    for i, label_text in enumerate(labels):
        Label(root, text=label_text).grid(row=i, column=0, padx=10, pady=5, sticky="e")
        if label_text == "Age":
            entry = Entry(root, validate="key", validatecommand=(root.register(validate_age), "%P"))
        else:
            entry = Entry(root)
        entry.grid(row=i, column=1, padx=10, pady=5, sticky="ew")
        entry.bind("<Return>", next_field)
        entries.append(entry)
    #extract the entry fields from the entries list and assign them to the global variables
    first_name_entry, last_name_entry, age_entry, email_entry, phone_entry = entries
    #Create a frame for the buttons and add the "Add" and "Clear" buttons to it
    button_frame = Frame(root)
    button_frame.grid(row=5, column=0, columnspan=2, pady=10)
    Button(button_frame, text="Add", width=12, command=add_data).pack(side="left", padx=5)
    Button(button_frame, text="Clear", width=12, command=clear_form).pack(side="left", padx=5)
    #Create a status label to display messages to the user
    status_label = Label(root, text="", anchor="w")
    status_label.grid(row=6, column=0, columnspan=2, padx=10, sticky="w")

    root.mainloop()


if __name__ == "__main__":
    init_excel()
    build_gui()