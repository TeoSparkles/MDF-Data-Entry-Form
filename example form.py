from pathlib import Path
from tkinter import END, Tk
from tkinter import ttk

from openpyxl import Workbook, load_workbook

EXCEL_FILE = Path(__file__).with_name("MDF-Data-Form.xlsx")

first_name_entry = None
last_name_entry = None
age_entry = None
email_entry = None
phone_entry = None
status_label = None
wb = None
ws = None


def init_excel():
    global wb, ws

    if EXCEL_FILE.exists():
        wb = load_workbook(EXCEL_FILE)
    else:
        wb = Workbook()

    ws = wb.active
    headers = ["First Name", "Last Name", "Age", "Email", "Phone"]

    if ws.max_row == 0 or ws.cell(row=1, column=1).value != headers[0]:
        for i, header in enumerate(headers, start=1):
            ws.cell(row=1, column=i).value = header

    wb.save(EXCEL_FILE)
    return wb, ws


def add_data():
    first_name = first_name_entry.get().strip()
    last_name = last_name_entry.get().strip()
    age = age_entry.get().strip()
    email = email_entry.get().strip()
    phone = phone_entry.get().strip()

    if first_name and last_name and age and email and phone:
        ws.append([first_name, last_name, age, email, phone])
        wb.save(EXCEL_FILE)
        first_name_entry.delete(0, END)
        last_name_entry.delete(0, END)
        age_entry.delete(0, END)
        email_entry.delete(0, END)
        phone_entry.delete(0, END)
        status_label.config(text="Data added successfully!", foreground="green")
    else:
        status_label.config(text="Please fill in all fields.", foreground="red")


def clear_form():
    first_name_entry.delete(0, END)
    last_name_entry.delete(0, END)
    age_entry.delete(0, END)
    email_entry.delete(0, END)
    phone_entry.delete(0, END)
    status_label.config(text="Form cleared.", foreground="blue")


def validate_age(new_value):
    if new_value == "":
        return True
    return new_value.isdigit() and len(new_value) <= 3


def next_field(event):
    if event.widget == phone_entry:
        add_data()
    else:
        event.widget.tk_focusNext().focus()


def build_gui():
    global first_name_entry, last_name_entry, age_entry, email_entry, phone_entry, status_label

    root = Tk()
    root.title("MDF Data Entry Form")
    root.geometry("430x330")
    root.resizable(False, False)
    root.configure(padx=20, pady=20)

    style = ttk.Style()
    style.theme_use("default")
    style.configure("TLabel", font=("Segoe UI", 10))
    style.configure("Header.TLabel", font=("Segoe UI", 11, "bold"))
    style.configure("TEntry", padding=6)
    style.configure("TButton", padding=(10, 6), font=("Segoe UI", 10))

    main_frame = ttk.Frame(root, padding=16)
    main_frame.grid(row=0, column=0, sticky="nsew")
    main_frame.columnconfigure(1, weight=1)

    ttk.Label(main_frame, text="Enter your information", style="Header.TLabel").grid(
        row=0, column=0, columnspan=2, pady=(0, 12), sticky="w"
    )

    labels = ["First Name", "Last Name", "Age", "Email", "Phone"]
    entries = []

    for i, label_text in enumerate(labels):
        ttk.Label(main_frame, text=label_text).grid(row=i + 1, column=0, padx=(0, 10), pady=6, sticky="e")
        if label_text == "Age":
            entry = ttk.Entry(main_frame, validate="key", validatecommand=(root.register(validate_age), "%P"), width=30)
        else:
            entry = ttk.Entry(main_frame, width=30)
        entry.grid(row=i + 1, column=1, pady=6, sticky="ew")
        entry.bind("<Return>", next_field)
        entries.append(entry)

    first_name_entry, last_name_entry, age_entry, email_entry, phone_entry = entries

    button_frame = ttk.Frame(main_frame)
    button_frame.grid(row=len(labels) + 1, column=0, columnspan=2, pady=(12, 8))
    ttk.Button(button_frame, text="Add", command=add_data).pack(side="left", padx=5)
    ttk.Button(button_frame, text="Clear", command=clear_form).pack(side="left", padx=5)

    status_label = ttk.Label(main_frame, text="", foreground="blue", wraplength=360)
    status_label.grid(row=len(labels) + 2, column=0, columnspan=2, sticky="w", pady=(6, 0))

    root.bind("<Escape>", lambda event: root.destroy())
    root.mainloop()


if __name__ == "__main__":
    init_excel()
    build_gui()